from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass
class ImportRow:
    row_num: int
    data: dict[str, Any]
    errors: list[str]


def _as_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _as_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    return None


def _read_sheet_bytes(xlsx_bytes: bytes, *, expected_headers: list[str]) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Read header row
    header = []
    for cell in ws[1]:
        header.append((cell.value or "").strip() if isinstance(cell.value, str) else (cell.value or ""))
    header = [str(h).strip().upper() for h in header if str(h).strip() != ""]

    file_errors: list[str] = []
    exp = [h.strip().upper() for h in expected_headers]
    if header != exp:
        file_errors.append(
            f"Invalid header. Expected: {', '.join(exp)}. Found: {', '.join(header) if header else '(blank)'}."
        )
        return [], file_errors

    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None:
            continue
        # consider empty row
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue
        row = {exp[i]: r[i] if i < len(r) else None for i in range(len(exp))}
        rows.append(row)
    return rows, file_errors


def _read_csv_bytes(csv_bytes: bytes, *, expected_headers: list[str]) -> tuple[list[dict[str, Any]], list[str]]:
    # Decode with UTF-8-SIG first (Excel), fallback to cp1252.
    try:
        text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = csv_bytes.decode("cp1252", errors="replace")

    file_errors: list[str] = []
    exp = [h.strip().upper() for h in expected_headers]

    reader = csv.DictReader(text.splitlines())
    if not reader.fieldnames:
        file_errors.append("Invalid header. File appears to be empty.")
        return [], file_errors

    found = [str(h or "").strip().upper() for h in reader.fieldnames]
    if found != exp:
        file_errors.append(
            f"Invalid header. Expected: {', '.join(exp)}. Found: {', '.join(found) if found else '(blank)'}."
        )
        return [], file_errors

    rows: list[dict[str, Any]] = []
    for r in reader:
        # skip completely blank rows
        if all((str(v or "").strip() == "") for v in r.values()):
            continue
        rows.append({k.strip().upper(): v for k, v in r.items()})
    return rows, []


def parse_fees_xlsx(xlsx_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_AMOUNT", "GST_AMOUNT"]
    raw_rows, file_errors = _read_sheet_bytes(xlsx_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r["DATE"])
        if not d:
            errors.append("DATE must be a valid date (e.g. 2026-05-10).")

        client_id = (str(r["CLIENT_ID"] or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r["CLIENT_NAME"] or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")

        fees = _as_decimal(r["FEES_AMOUNT"])
        gst = _as_decimal(r["GST_AMOUNT"]) if r.get("GST_AMOUNT") is not None else Decimal("0")
        if fees is None:
            errors.append("FEES_AMOUNT must be a number.")
        elif fees < 0:
            errors.append("FEES_AMOUNT cannot be negative.")
        if gst is None:
            errors.append("GST_AMOUNT must be a number (or blank for 0).")
        elif gst < 0:
            errors.append("GST_AMOUNT cannot be negative.")

        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_amount": fees,
                    "gst_amount": gst or Decimal("0.00"),
                },
                errors=errors,
            )
        )
    return out, []


def parse_fees_csv(csv_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_AMOUNT", "GST_AMOUNT"]
    raw_rows, file_errors = _read_csv_bytes(csv_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r.get("DATE"))
        if not d:
            errors.append("DATE must be a valid date (e.g. 2026-05-10).")

        client_id = (str(r.get("CLIENT_ID") or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r.get("CLIENT_NAME") or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")

        fees = _as_decimal(r.get("FEES_AMOUNT"))
        gst = _as_decimal(r.get("GST_AMOUNT")) if (r.get("GST_AMOUNT") not in (None, "")) else Decimal("0.00")
        if fees is None:
            errors.append("FEES_AMOUNT must be a number.")
        elif fees < 0:
            errors.append("FEES_AMOUNT cannot be negative.")
        if gst is None:
            errors.append("GST_AMOUNT must be a number (or blank for 0).")
        elif gst < 0:
            errors.append("GST_AMOUNT cannot be negative.")

        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_amount": fees,
                    "gst_amount": gst,
                },
                errors=errors,
            )
        )
    return out, []


def _parse_receipt_received_amounts(row: dict, errors: list[str]) -> tuple[Decimal | None, Decimal | None]:
    """Parse fees_received / expenses_received; RECEIPTS_AMOUNT (legacy) maps to fees only."""
    legacy = row.get("RECEIPTS_AMOUNT")
    fees_raw = row.get("FEES_RECEIVED_AMOUNT")
    exp_raw = row.get("EXPENSES_RECEIVED_AMOUNT")
    fees = _as_decimal(fees_raw) if fees_raw not in (None, "") else None
    if fees is None and legacy not in (None, ""):
        fees = _as_decimal(legacy)
    expenses = _as_decimal(exp_raw) if exp_raw not in (None, "") else None
    for label, val in (
        ("FEES_RECEIVED_AMOUNT", fees_raw, fees),
        ("EXPENSES_RECEIVED_AMOUNT", exp_raw, expenses),
        ("RECEIPTS_AMOUNT", legacy, fees if legacy not in (None, "") else None),
    ):
        if val is not None and val < 0:
            errors.append(f"{label} cannot be negative.")
    if fees is None and expenses is None:
        if any(x not in (None, "") for x in (fees_raw, exp_raw, legacy)):
            errors.append("Fees received and/or expenses received must be valid numbers.")
        else:
            errors.append("Enter FEES_RECEIVED_AMOUNT and/or EXPENSES_RECEIVED_AMOUNT (or legacy RECEIPTS_AMOUNT for fees).")
    elif (fees or Decimal("0")) <= 0 and (expenses or Decimal("0")) <= 0:
        errors.append("At least one of fees received or expenses received must be greater than zero.")
    return fees or Decimal("0"), expenses or Decimal("0")


def parse_receipts_xlsx(xlsx_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_RECEIVED_AMOUNT", "EXPENSES_RECEIVED_AMOUNT"]
    raw_rows, file_errors = _read_sheet_bytes(xlsx_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r["DATE"])
        if not d:
            errors.append("DATE must be a valid date.")
        client_id = (str(r["CLIENT_ID"] or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r["CLIENT_NAME"] or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")
        fees_received, expenses_received = _parse_receipt_received_amounts(r, errors)
        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_received": fees_received,
                    "expenses_received": expenses_received,
                },
                errors=errors,
            )
        )
    return out, []


def parse_receipts_csv(csv_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_RECEIVED_AMOUNT", "EXPENSES_RECEIVED_AMOUNT"]
    raw_rows, file_errors = _read_csv_bytes(csv_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r.get("DATE"))
        if not d:
            errors.append("DATE must be a valid date.")
        client_id = (str(r.get("CLIENT_ID") or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r.get("CLIENT_NAME") or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")
        fees_received, expenses_received = _parse_receipt_received_amounts(r, errors)
        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_received": fees_received,
                    "expenses_received": expenses_received,
                },
                errors=errors,
            )
        )
    return out, []


def parse_expenses_xlsx(xlsx_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_AMOUNT", "EXPENSES_AMOUNT"]
    raw_rows, file_errors = _read_sheet_bytes(xlsx_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r["DATE"])
        if not d:
            errors.append("DATE must be a valid date.")
        client_id = (str(r["CLIENT_ID"] or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r["CLIENT_NAME"] or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")
        fee_amt = _as_decimal(r["FEES_AMOUNT"])
        if fee_amt is None:
            errors.append("FEES_AMOUNT must be a number.")
        elif fee_amt < 0:
            errors.append("FEES_AMOUNT cannot be negative.")
        amt = _as_decimal(r["EXPENSES_AMOUNT"])
        if amt is None:
            errors.append("EXPENSES_AMOUNT must be a number.")
        elif amt < 0:
            errors.append("EXPENSES_AMOUNT cannot be negative.")
        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_amount": fee_amt,
                    "expenses_paid": amt,
                },
                errors=errors,
            )
        )
    return out, []


def parse_expenses_csv(csv_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    expected = ["DATE", "CLIENT_ID", "CLIENT_NAME", "FEES_AMOUNT", "EXPENSES_AMOUNT"]
    raw_rows, file_errors = _read_csv_bytes(csv_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r.get("DATE"))
        if not d:
            errors.append("DATE must be a valid date.")
        client_id = (str(r.get("CLIENT_ID") or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r.get("CLIENT_NAME") or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")
        fee_amt = _as_decimal(r.get("FEES_AMOUNT"))
        if fee_amt is None:
            errors.append("FEES_AMOUNT must be a number.")
        elif fee_amt < 0:
            errors.append("FEES_AMOUNT cannot be negative.")
        amt = _as_decimal(r.get("EXPENSES_AMOUNT"))
        if amt is None:
            errors.append("EXPENSES_AMOUNT must be a number.")
        elif amt < 0:
            errors.append("EXPENSES_AMOUNT cannot be negative.")
        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_amount": fee_amt,
                    "expenses_paid": amt,
                },
                errors=errors,
            )
        )
    return out, []


def parse_mis_combined_csv(csv_bytes: bytes) -> tuple[list[ImportRow], list[str]]:
    """
    Combined MIS bulk upload.

    Header:
      DATE, CLIENT_ID, CLIENT_NAME, FEES_AMOUNT, GST_AMOUNT, FEES_RECEIVED_AMOUNT,
      EXPENSES_RECEIVED_AMOUNT, EXPENSES_AMOUNT
    """
    expected = [
        "DATE",
        "CLIENT_ID",
        "CLIENT_NAME",
        "FEES_AMOUNT",
        "GST_AMOUNT",
        "FEES_RECEIVED_AMOUNT",
        "EXPENSES_RECEIVED_AMOUNT",
        "EXPENSES_AMOUNT",
    ]
    raw_rows, file_errors = _read_csv_bytes(csv_bytes, expected_headers=expected)
    out: list[ImportRow] = []
    if file_errors:
        return out, file_errors

    for idx, r in enumerate(raw_rows, start=2):
        errors: list[str] = []
        d = _as_date(r.get("DATE"))
        if not d:
            errors.append("DATE must be a valid date (e.g. 2026-05-10).")

        client_id = (str(r.get("CLIENT_ID") or "").strip().upper())
        if not client_id:
            errors.append("CLIENT_ID is required.")
        client_name = (str(r.get("CLIENT_NAME") or "").strip())
        if not client_name:
            errors.append("CLIENT_NAME is required.")

        fees = _as_decimal(r.get("FEES_AMOUNT")) if (r.get("FEES_AMOUNT") not in (None, "")) else None
        gst = _as_decimal(r.get("GST_AMOUNT")) if (r.get("GST_AMOUNT") not in (None, "")) else None
        fees_received = (
            _as_decimal(r.get("FEES_RECEIVED_AMOUNT"))
            if (r.get("FEES_RECEIVED_AMOUNT") not in (None, ""))
            else None
        )
        if fees_received is None and (r.get("RECEIPTS_AMOUNT") not in (None, "")):
            fees_received = _as_decimal(r.get("RECEIPTS_AMOUNT"))
        expenses_received = (
            _as_decimal(r.get("EXPENSES_RECEIVED_AMOUNT"))
            if (r.get("EXPENSES_RECEIVED_AMOUNT") not in (None, ""))
            else None
        )
        expenses = (
            _as_decimal(r.get("EXPENSES_AMOUNT")) if (r.get("EXPENSES_AMOUNT") not in (None, "")) else None
        )

        for label, raw_val, val in (
            ("FEES_AMOUNT", r.get("FEES_AMOUNT"), fees),
            ("GST_AMOUNT", r.get("GST_AMOUNT"), gst),
            ("FEES_RECEIVED_AMOUNT", r.get("FEES_RECEIVED_AMOUNT"), fees_received),
            ("EXPENSES_RECEIVED_AMOUNT", r.get("EXPENSES_RECEIVED_AMOUNT"), expenses_received),
            ("RECEIPTS_AMOUNT", r.get("RECEIPTS_AMOUNT"), fees_received if r.get("RECEIPTS_AMOUNT") not in (None, "") else None),
            ("EXPENSES_AMOUNT", r.get("EXPENSES_AMOUNT"), expenses),
        ):
            if raw_val not in (None, "") and val is None:
                errors.append(f"{label} must be a number (or blank).")
            if val is not None and val < 0:
                errors.append(f"{label} cannot be negative.")

        # GST rule: GST cannot be entered when Fees is 0/blank.
        if gst is not None and gst > Decimal("0") and (fees is None or fees == Decimal("0")):
            errors.append("GST_AMOUNT cannot be entered when FEES_AMOUNT is 0/blank.")

        # Only create records where amount exists and is > 0 (blank => ignore).
        has_any = any(
            v is not None and v != Decimal("0")
            for v in (fees, gst, fees_received, expenses_received, expenses)
        )
        if not has_any:
            errors.append(
                "Provide at least one amount (fees/gst/fees received/expenses received/expenses paid) or remove the row."
            )

        out.append(
            ImportRow(
                row_num=idx,
                data={
                    "date": d,
                    "client_id": client_id,
                    "client_name": client_name,
                    "fees_amount": fees,
                    "gst_amount": gst,
                    "fees_received_amount": fees_received,
                    "expenses_received_amount": expenses_received,
                    "expenses_amount": expenses,
                },
                errors=errors,
            )
        )
    return out, []

